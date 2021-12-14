import openpyxl as xl


wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet.cell(1, 1)

for row in range(3, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value + 30
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price


wb.save('transactions1.xlsx')
