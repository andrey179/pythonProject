import time
from datetime import datetime
import openpyxl as xl
import pywhatkit

wb = xl.load_workbook('PhoneData.xlsx')
sheet = wb['Sheet1']  # open workbook


# sheet , row, col,
def get_cell_value(sheet, row, col):
    cell = sheet.cell(row, col)
    print(cell.value)
    if col == 5 or col == 6:
        return int(cell.value)
    else:
        return str(cell.value)


def get_cell(sheet, row, col):
    cell = sheet.cell(row, col)
    print(cell.value)
    return cell


today_date = f"{datetime.now().year}-0{datetime.now().month}-0{datetime.now().day}"


def check_date():
    for row in range(2, sheet.max_row + 1):
        get_date = str(get_cell_value(sheet, row, 4))
        corrected_date = get_date[0:10]
        if corrected_date == today_date:
            DataComponents()
        else:
            print("Not today")
            pass


class DataComponents:
    for row in range(2, sheet.max_row + 1):
        get_phone = "+972"
        get_phone += str(get_cell_value(sheet, row, 2))  # Getting phone number as a string
        get_message = get_cell_value(sheet, row, 3)  # Getting message
        get_hours = get_cell_value(sheet, row, 5)  # Getting hours
        get_minutes = get_cell_value(sheet, row, 6)  # Getting minutes
        last_date = get_cell(sheet, row, 7)
        update_date = f"{datetime.now().day}-{datetime.now().month} {datetime.now().hour}:{datetime.now().minute}"
        last_date.value = update_date  # Overwrite to current data
        pywhatkit.sendwhatmsg(get_phone, get_message, get_hours, get_minutes)
        time.sleep(10)
        pass


wb.save('PhoneData.xlsx')  # save
